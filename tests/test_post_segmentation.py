"""Tests for src.post_segmentation."""

import importlib
import io
import sys
import tempfile
from pathlib import Path

import numpy as np
import pandas as pd
import pytest
from openpyxl import load_workbook

from src.post_segmentation import PostSegConfig, run
from src.post_segmentation.analyze import (
    _apply_filters,
    _graphpad_long,
    _ordered_columns,
    _per_group_tabs,
    _validate_columns,
)
from src.post_segmentation.stats import (
    add_flag_robust_outlier_mad,
    group_descriptive_stats,
    median_absolute_deviation,
)


def _sample_metrics(extra_cols=True):
    """A minimal metrics CSV-like DataFrame with two groups."""
    base = {
        "object_id": [1, 2, 3, 4, 5, 6],
        "image": ["A.tif", "A.tif", "A.tif", "B.tif", "B.tif", "B.tif"],
        "area_px": [12000, 30000, 600000, 15000, 25000, 45000],
        "area_um2": [120.0, 300.0, 6000.0, 150.0, 250.0, 450.0],
        "circularity": [0.7, 0.6, 0.2, 0.8, 0.4, 0.9],
        "equivalent_diameter_um": [12.0, 18.0, 80.0, 14.0, 16.0, 22.0],
    }
    if extra_cols:
        base["condition"] = ["ctrl", "ctrl", "ctrl", "drug", "drug", "drug"]
    return pd.DataFrame(base)


# ---------- stats.py ----------


def test_mad_simple():
    assert median_absolute_deviation(np.array([1, 1, 2, 2, 4])) == 1.0


def test_mad_empty_or_singleton():
    assert median_absolute_deviation(np.array([])) == 0.0
    assert median_absolute_deviation(np.array([5.0])) == 0.0


def test_add_flag_robust_outlier_mad():
    df = _sample_metrics()
    out = add_flag_robust_outlier_mad(df, group_cols=["condition"], metric_col="area_px")
    # ctrl: areas [12000, 30000, 600000], median=30000, abs-deviations=[18000, 0, 570000],
    # MAD = median([0, 18000, 570000]) = 18000
    # flags = [1.0, 0.0, 31.666...]
    ctrl = out[out["condition"] == "ctrl"]["flag_robust_outlier_mad"].tolist()
    assert ctrl[0] == pytest.approx(1.0)
    assert ctrl[1] == pytest.approx(0.0)
    assert ctrl[2] == pytest.approx(570000 / 18000)


def test_add_flag_zero_mad_group():
    """When MAD is 0 (all equal), flags should be 0 (no spread)."""
    df = pd.DataFrame({
        "condition": ["x", "x", "x"],
        "area_px": [100, 100, 100],
    })
    out = add_flag_robust_outlier_mad(df, ["condition"], "area_px")
    assert (out["flag_robust_outlier_mad"] == 0.0).all()


def test_group_descriptive_stats():
    df = _sample_metrics()
    stats = group_descriptive_stats(df, ["condition"], ["area_px", "circularity"])
    assert set(stats.columns) >= {
        "condition", "n", "area_px_mean", "area_px_median", "area_px_iqr", "area_px_mad",
        "circularity_mean", "circularity_median", "circularity_iqr", "circularity_mad",
    }
    drug_row = stats[stats["condition"] == "drug"].iloc[0]
    assert drug_row["n"] == 3
    assert drug_row["area_px_median"] == pytest.approx(25000)


# ---------- analyze.py ----------


def test_validate_columns_passes_when_present():
    df = _sample_metrics()
    _validate_columns(df, ["condition", "image"])  # should not raise


def test_validate_columns_raises_with_helpful_message():
    df = _sample_metrics()
    with pytest.raises(ValueError) as excinfo:
        _validate_columns(df, ["nonexistent"])
    msg = str(excinfo.value)
    assert "nonexistent" in msg
    assert "Parse filename into fields" in msg
    assert "SegOid GUI" in msg


def test_apply_filters_passed_filter_column():
    df = _sample_metrics()
    cfg = PostSegConfig(
        metrics_csv=Path("x"), output_xlsx=Path("y"),
        circularity_min=0.5, area_min_px=10000, area_max_px=50000,
    )
    out = _apply_filters(df, cfg)
    assert "passed_filter" in out.columns
    expected = [True, True, False, True, False, True]
    assert out["passed_filter"].tolist() == expected


def test_apply_filters_none_means_off():
    df = _sample_metrics()
    cfg = PostSegConfig(metrics_csv=Path("x"), output_xlsx=Path("y"))
    out = _apply_filters(df, cfg)
    # No filters configured → everything passes.
    assert out["passed_filter"].all()


def test_ordered_columns_priority():
    df = _sample_metrics()
    df["mean_intensity_green"] = 1.0
    df["mean_intensity_red"] = 2.0
    cols = _ordered_columns(df, ["condition"])
    # object_id and image come first
    assert cols[0] == "object_id"
    assert cols[1] == "image"
    # condition next
    assert "condition" in cols[:5]
    # intensities sorted alphabetically and grouped
    g_idx = cols.index("mean_intensity_green")
    r_idx = cols.index("mean_intensity_red")
    assert g_idx < r_idx


def test_per_group_tabs_only_passed_rows():
    df = _sample_metrics()
    df["passed_filter"] = [True, True, False, True, False, True]
    tabs = _per_group_tabs(df, ["condition"], list(df.columns))
    assert "ctrl" in tabs and "drug" in tabs
    assert len(tabs["ctrl"]) == 2  # one row dropped
    assert len(tabs["drug"]) == 2


def test_per_group_tabs_empty_when_all_filtered():
    """When every row in a group fails the filter, the tab still appears
    (with header columns only) so the user can see the group existed."""
    df = _sample_metrics()
    df["passed_filter"] = False  # nothing passes
    tabs = _per_group_tabs(df, ["condition"], list(df.columns))
    assert set(tabs.keys()) == {"ctrl", "drug"}
    assert len(tabs["ctrl"]) == 0
    assert len(tabs["drug"]) == 0
    # Columns present so the user sees the schema even with no rows
    assert "object_id" in tabs["ctrl"].columns
    assert "circularity" in tabs["drug"].columns


def test_per_group_tabs_partial_filter_one_group_empty():
    """Mixed: ctrl group has rows passing, drug group is fully filtered out.
    Both tabs should appear; drug should be empty."""
    df = _sample_metrics()
    df["passed_filter"] = [True, True, True, False, False, False]
    tabs = _per_group_tabs(df, ["condition"], list(df.columns))
    assert set(tabs.keys()) == {"ctrl", "drug"}
    assert len(tabs["ctrl"]) == 3
    assert len(tabs["drug"]) == 0


def test_graphpad_long_no_duplicate_object_id():
    df = _sample_metrics()
    df["passed_filter"] = True
    long = _graphpad_long(df, ["condition"])
    cols = list(long.columns)
    # 'object_id' should appear exactly once
    assert cols.count("object_id") == 1
    # 'group' should be first
    assert cols[0] == "group"


# ---------- end-to-end run() ----------


@pytest.fixture
def tmp_metrics_csv(tmp_path):
    df = _sample_metrics()
    csv = tmp_path / "metrics.csv"
    df.to_csv(csv, index=False)
    return csv


def test_run_writes_expected_sheets(tmp_path, tmp_metrics_csv):
    out = tmp_path / "post_seg.xlsx"
    cfg = PostSegConfig(
        metrics_csv=tmp_metrics_csv, output_xlsx=out, group_by=["condition"],
        circularity_min=0.5, area_min_px=10000, area_max_px=50000,
    )
    run(cfg)
    wb = load_workbook(out)
    assert set(wb.sheetnames) >= {
        "ctrl", "drug", "all_with_filter_status", "group_stats", "graphpad_long",
    }


def test_run_no_filename_reparse(tmp_metrics_csv, tmp_path):
    """Confirm post_segmentation does not import filename_schema parsing."""
    # Force a fresh import of the analyze module so we can inspect imports.
    if "src.post_segmentation.analyze" in sys.modules:
        importlib.reload(sys.modules["src.post_segmentation.analyze"])
    import src.post_segmentation.analyze as analyze_mod
    src = Path(analyze_mod.__file__).read_text()
    assert "parse_filename" not in src
    assert "filename_schema" not in src


def test_run_missing_group_column_errors(tmp_path, tmp_metrics_csv):
    out = tmp_path / "post_seg.xlsx"
    cfg = PostSegConfig(
        metrics_csv=tmp_metrics_csv, output_xlsx=out, group_by=["nonexistent"],
    )
    with pytest.raises(ValueError) as excinfo:
        run(cfg)
    assert "Parse filename into fields" in str(excinfo.value)


def test_run_default_group_by_is_image(tmp_path, tmp_metrics_csv):
    out = tmp_path / "post_seg.xlsx"
    cfg = PostSegConfig(metrics_csv=tmp_metrics_csv, output_xlsx=out)
    run(cfg)
    wb = load_workbook(out)
    # default group_by=['image'] → tabs named 'A.tif', 'B.tif' (sheet-safe)
    sheet_names = wb.sheetnames
    assert any("A" in s for s in sheet_names)
    assert any("B" in s for s in sheet_names)


def test_run_prints_pass_rate(tmp_path, tmp_metrics_csv, capsys):
    """The run summary should always print n/total passed."""
    out = tmp_path / "post_seg.xlsx"
    cfg = PostSegConfig(metrics_csv=tmp_metrics_csv, output_xlsx=out)
    run(cfg)
    captured = capsys.readouterr()
    # No filters → all 6 sample rows pass
    assert "Filter: 6/6 rows passed." in captured.out


def test_run_warns_when_all_filtered(tmp_path, tmp_metrics_csv, capsys):
    """When all rows are excluded, print a loud warning pointing the user
    at all_with_filter_status, so they don't quietly get empty tabs."""
    out = tmp_path / "post_seg.xlsx"
    cfg = PostSegConfig(
        metrics_csv=tmp_metrics_csv, output_xlsx=out, group_by=["condition"],
        circularity_min=10.0,  # impossible — circularity is 0..1
    )
    run(cfg)
    captured = capsys.readouterr()
    assert "Filter: 0/6 rows passed." in captured.out
    assert "WARNING" in captured.out
    assert "all_with_filter_status" in captured.out

    # And the per-group tabs still exist (empty), not silently dropped.
    wb = load_workbook(out)
    assert "ctrl" in wb.sheetnames
    assert "drug" in wb.sheetnames
    ctrl = pd.read_excel(out, sheet_name="ctrl")
    assert len(ctrl) == 0
    assert "object_id" in ctrl.columns


def test_excel_writer_safe_sheet_name(tmp_path):
    """Sheet names with forbidden chars get sanitized."""
    from src.post_segmentation.excel_writer import write_multitab_excel
    out = tmp_path / "out.xlsx"
    sheets = {
        "name/with/slashes": pd.DataFrame({"a": [1]}),
        "x" * 50: pd.DataFrame({"a": [1]}),  # too long
    }
    write_multitab_excel(out, sheets)
    wb = load_workbook(out)
    for name in wb.sheetnames:
        assert "/" not in name
        assert len(name) <= 31
